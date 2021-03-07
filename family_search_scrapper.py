import time

import sys
sys.stdin.reconfigure(encoding='utf-8')
sys.stdout.reconfigure(encoding='utf-8')

import re
import requests

from fs_parser.pr_parser import PRParser

# from write_excel import WriteExcel
from openpyxl import load_workbook, Workbook

class FamilySearchScrapper(PRParser):
    def __init__(self, **kwargs):
        # self.wx = WriteExcel()
        self.count = 0
        self.filename = kwargs.get('filename', None)
        self.id = kwargs.get('id', None) # id for writing logs

        if (self.filename is None):
            raise Exception('Please provide a filename')

        wb = Workbook()
        wb.save(self.filename)
        self.match = kwargs.get('match', 'exact')
        self.firstname = kwargs.get('firstname', '').replace(' ', '%20')
        self.surname = kwargs.get('surname', '').replace(' ', '%20')

        self.req = requests.Session()
        print("Logging in...")
        login_url = 'https://www.familysearch.org/auth/familysearch/login?fhf=true&returnUrl=%2F'
        html = self.req.get(login_url)
        params = re.findall(r'<input type=\"hidden\" name=\"params\" value=\"([\w\-\=]+)\"\/>', html.text)[0]

        # print(params)

        headers = {
            'Content-Type': 'application/x-www-form-urlencoded',
            'Connection': 'keep-alive',
            'Origin': 'https://ident.familysearch.org',
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Referer': html.url,
        }

        body = {
            'userName': 'ww1_project',
            'password': 'maledizione222',
            'params': params
        }

        self.req.post('https://ident.familysearch.org/cis-web/oauth2/v3/authorization', data=body, headers=headers)
        # print(post.text)

    
    def search(self):
        if (self.firstname == '' and self.surname == ''):
            raise Exception("Must have firstname or surname")

        if self.match == 'approximate':
            self.search_url = f"https://www.familysearch.org/service/search/hr/v2/personas?q.givenName={self.firstname}&q.surname={self.surname}&f.collectionId=2125045&count=20&offset=0&m.defaultFacets=on&m.queryRequireDefault=on&m.facetNestCollectionInCategory=on"
        elif self.match == "exact":
            self.search_url = f"https://www.familysearch.org/service/search/hr/v2/personas?q.givenName={self.firstname}&q.givenName.exact=on&q.surname={self.surname}&q.surname.exact=on&f.collectionId=2125045&count=20&offset=0&m.defaultFacets=on&m.queryRequireDefault=on&m.facetNestCollectionInCategory=on"
        else:
            raise Exception("Unknown match criteria")

        # start = time.time()

        test = self.req.get(self.search_url)
        source = test.text

        links = re.findall(r'https://www.familysearch.org/ark:/[0-9]+/[0-9]+:[0-9]+:[\w-]+', source)
        if (len(links) == 0):
            print(f"Not found for {self.firstname.replace('%20',' ')}")
            log_file = open(f'log/logs.txt', 'a+')
            log_file.write(f'[{self.id}]     Status: Failed to fetch\n')
            log_file.close()
            return

        
        fssessionid = self.req.cookies.get_dict()['fssessionid']


        """
            Sample: 
                :authority: www.familysearch.org
                :method: GET
                :path: /ark:/61903/1:1:QVBP-VJ4W
                :scheme: https
                accept: application/x-gedcomx-v1+json, application/json
                accept-encoding: gzip, deflate, br
                accept-language: en
                authorization: Bearer 907c9295-69ad-4841-85c5-e5d18992ba4c-prod
        """
        data_header = {
            'Accept' : 'application/x-gedcomx-v1+json, application/json',
            'Accept-Encoding' : 'gzip, deflate, br',
            'Accept-Language' : 'en',
            'Authorization': 'Bearer ' + fssessionid,
            'Referer' : self.search_url,
            'from' : 'fsSearch.record.getGedcomX@familysearch.org',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin'
        }
        for link in links[0:1]:
            start = time.time()
            data = self.req.get(links[0], headers = data_header)
            super().__init__(data.text)

            wb = load_workbook(self.filename)
            # Select First Worksheet
            ws = wb.worksheets[0]
            # Append 2 new Rows - Columns A - D
            ws.append(super().get_data())
            print(self.get_data())
            wb.save(self.filename)
            self.count += 1

            end = time.time()
            print(end - start)

    def get_data(self):
        return super().get_data()
    
    def clear_data(self):
        super().clear_data()

if __name__ == '__main__':
    obj = FamilySearchScrapper()
    obj.firstname = 'A'
    obj.surname = 'Thomas'
    obj.search()
    print(obj.get_data())