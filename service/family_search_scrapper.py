import time

import sys
sys.stdin.reconfigure(encoding='utf-8')
sys.stdout.reconfigure(encoding='utf-8')

import re
import requests

from service.fs_parser.pr_parser import PRParser

# from write_excel import WriteExcel
from openpyxl import load_workbook, Workbook

class FamilySearchScrapper(PRParser):
    def __init__(self, **kwargs):
        self.read_credentials()
        self.count = 0
        self.filename = kwargs.get('filename', None)
        self.id = kwargs.get('id', None) # id for writing logs

        if (self.filename is None):
            raise Exception('Please provide a filename')


        """                                     Create an excel file if not exists or open to append more row                       """
        try:
            # open existing excel file to append
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.worksheets[0]
        except FileNotFoundError:
            # create new excel file
            self.wb = Workbook()
            self.ws = self.wb.worksheets[0]
            headers = ['Name', 'Event Type', 'Event Date', 'Event Year', 'Residence Place', 'Age', 'Military Company/Regiment', 'Military Regiment', 'Batallion', 'Birth Year', 'Birthplace']
            self.ws.append(headers)
            self.wb.save(self.filename)


        self.match = kwargs.get('match', 'exact')
        self.firstname = kwargs.get('firstname', '').replace(' ', '%20')
        self.surname = kwargs.get('surname', '').replace(' ', '%20')

        self.req = requests.Session()
        print("Logging in...")
        login_url = 'https://www.familysearch.org/auth/familysearch/login?fhf=true&returnUrl=%2F'
        html = self.req.get(login_url)
        params = re.findall(r'<input type=\"hidden\" name=\"params\" value=\"([\w\-\=]+)\"\/>', html.text)[0]

        # print(params)

        headers = {
            'Content-Type': 'application/x-www-form-urlencoded',
            'Connection': 'keep-alive',
            'Origin': 'https://ident.familysearch.org',
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Referer': html.url,
        }

        body = {
            'userName': self.username,
            'password': self.password,
            'params': params
        }

        self.req.post('https://ident.familysearch.org/cis-web/oauth2/v3/authorization', data=body, headers=headers)
        # print(post.text)

    
    def search(self):
        if (self.firstname == '' and self.surname == ''):
            raise Exception("Must have firstname or surname")

        
        if self.match == 'approximate':
            self.search_url = f"https://www.familysearch.org/service/search/hr/v2/personas?q.givenName={self.firstname}&q.surname={self.surname}&f.collectionId={self.collection_id}&count=100&offset=0&m.defaultFacets=on&m.queryRequireDefault=on&m.facetNestCollectionInCategory=on"
        elif self.match == "exact":
            self.search_url = f"https://www.familysearch.org/service/search/hr/v2/personas?q.givenName={self.firstname}&q.givenName.exact=on&q.surname={self.surname}&q.surname.exact=on&f.collectionId={self.collection_id}&count=100&offset=0&m.defaultFacets=on&m.queryRequireDefault=on&m.facetNestCollectionInCategory=on"
        else:
            raise Exception("Unknown match criteria")

        # start = time.time()

        test = self.req.get(self.search_url)
        source = test.text

        links = re.findall(r'https://www.familysearch.org/ark:/[0-9]+/[0-9]+:[0-9]+:[\w-]+', source)
        if (len(links) == 0):
            print(f"Not found for {self.firstname.replace('%20',' ')} {self.surname.replace('%20',' ')}")
            log_file = open(f'log/logs.txt', 'a+')
            log_file.write(f'[{self.id}]     Status: Failed to fetch\n')
            log_file.close()
            return

        
        fssessionid = self.req.cookies.get_dict()['fssessionid']


        """
            Sample: 
                :authority: www.familysearch.org
                :method: GET
                :path: /ark:/61903/1:1:QVBP-VJ4W
                :scheme: https
                accept: application/x-gedcomx-v1+json, application/json
                accept-encoding: gzip, deflate, br
                accept-language: en
                authorization: Bearer 907c9295-69ad-4841-85c5-e5d18992ba4c-prod
        """
        data_header = {
            'Accept' : 'application/x-gedcomx-v1+json, application/json',
            'Accept-Encoding' : 'gzip, deflate, br',
            'Accept-Language' : 'en',
            'Authorization': 'Bearer ' + fssessionid,
            'Referer' : self.search_url,
            'from' : 'fsSearch.record.getGedcomX@familysearch.org',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin'
        }
        links = set(links)
        for link in links:
            # start = time.time()           # <----- start timer
            data = self.req.get(link, headers = data_header)
            super().__init__(data.text)

            # data_list = super().get_data()
            # for i in range(1, 12):
            #     self.ws.cell(column = i, row = self.count, value = data_list[i-1])
            # ws.append(super().get_data())

            self.ws.append(super().get_data())
            print(f'{self.get_data()[0]} has been written')
            self.count += 1
            self.wb.save(self.filename)

            # end = time.time()            # <--------- end timer
            # print(end - start)

    def read_existing_excel(self):
        for i in range(2, self.count+2):
            for j in range(1, 12):
                self.ws.cell(column=j, row=i, value=self.ws.cell(column=j, row=i).value)
                
        self.count += 2
    def get_data(self):
        return super().get_data()
    
    def clear_data(self):
        super().clear_data()
    
    def read_credentials(self):
        cred_file = open('credentials.txt', 'r')
        content = cred_file.readlines()
        self.username = content[0].strip()
        self.password = content[1].strip()
        self.collection_id = content[2].strip()
        cred_file.close()

if __name__ == '__main__':
    obj = FamilySearchScrapper(filename = 'temp.xlsx', count = 0)
    obj.firstname = 'Thomas W'
    obj.surname = 'Craggs'
    obj.search()
    # print(obj.get_data())